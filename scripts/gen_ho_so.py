from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

BLACK     = "FF000000"
WHITE     = "FFFFFFFF"
GOLD      = "FFC9A84C"
GOLD_LIGHT= "FFFFF3CD"
DARK_BG   = "FF1C1C1C"
MID_BG    = "FF2E2E2E"
GREY_ROW  = "FFF5F5F5"
FONT_MAIN = "Arial"

def hdr_font(sz=11, bold=True, color=WHITE):
    return Font(name=FONT_MAIN, size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color=BLACK):
    return Font(name=FONT_MAIN, size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

thin = Side(style="thin", color="FFD0D0D0")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def section_header(ws, row, col, text, width_cols=2):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+width_cols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font    = hdr_font(11, True, WHITE)
    c.fill    = fill(GOLD)
    c.alignment = left(wrap=False)
    c.border  = thin_border()

def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = body_font(10, bold=True, color="FF444444")
    c.fill      = fill("FFF0EBD8")
    c.alignment = left(wrap=False)
    c.border    = thin_border()

def value_cell(ws, row, col, text, row_shade=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = body_font(10, color=BLACK)
    c.fill      = fill(GREY_ROW if row_shade else WHITE)
    c.alignment = left(wrap=True)
    c.border    = thin_border()

# ── Sheet 1: Tong Quan ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Tong Quan"
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 14

ws1.merge_cells("A1:C1")
c = ws1["A1"]
c.value     = "MAISON DENUDE - HO SO DOANH NGHIEP"
c.font      = Font(name=FONT_MAIN, size=16, bold=True, color=GOLD)
c.fill      = fill(DARK_BG)
c.alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:C2")
c = ws1["A2"]
c.value     = "Tai lieu noi bo - cap nhat 22/06/2026"
c.font      = Font(name=FONT_MAIN, size=9, italic=True, color="FF888888")
c.fill      = fill(MID_BG)
c.alignment = center()
ws1.row_dimensions[2].height = 18

row = 4
section_header(ws1, row, 1, "THONG TIN CO BAN", 3); row += 1
fields_basic = [
    ("Ten thuong hieu",   "Maison Denude"),
    ("Loai hinh",          "Thoi trang bespoke & artisanal cao cap (designer house, KHONG phai tiem may)"),
    ("Dia chi",            "194 Le Thanh Ton, Quan 1, TP.HCM (tang 2)"),
    ("Nguoi sang lap",     "Chi Chi (Chi Bui) - fashion designer"),
    ("Dinh vi gia",        "Premium / Upscale (Le Le Le)"),
    ("San pham chinh",     "Ao dai bespoke, trang phuc dip dac biet, dam thiet ke, ao tam (swimwear)"),
    ("Facebook followers", "~4.400"),
    ("Slogan dinh vi",     "Made in Saigon - by a Vietnamese designer"),
]
for i, (lbl, val) in enumerate(fields_basic):
    label_cell(ws1, row, 1, lbl)
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws1, row, 2, val, row_shade=(i % 2 == 0))
    ws1.row_dimensions[row].height = 22
    row += 1

row += 1
section_header(ws1, row, 1, "DINH VI & GIA TRI COT LOI", 3); row += 1
fields_pos = [
    ("Core Offer",         "Thiet ke theo yeu cau (bespoke) - khong san xuat dai tra; moi san pham duoc lam rieng cho tung khach"),
    ("Phong cach",         "Heritage-inspired x sultry x feminine power - giao thoa truyen thong Viet va phong cach hien dai"),
    ("Diem khac biet",     "Thu cong cao cap: hand embroidery, hand beading, embroidery art; KHONG phai tiem may"),
    ("Doi tuong muc tieu", "Phu nu muon mac trang phuc doc nhat; khach quoc te tim nha thiet ke Viet Nam"),
    ("Tiem nang quoc te",  "Xu huong artisanal & slow fashion toan cau; Southeast Asian designers dang duoc chu y"),
    ("Ngon ngu thuong hieu","DUNG: premium, upscale, bespoke, artisanal. KHONG DUNG: 'high end', 'tailor', 'may do'"),
]
for i, (lbl, val) in enumerate(fields_pos):
    label_cell(ws1, row, 1, lbl)
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws1, row, 2, val, row_shade=(i % 2 == 0))
    ws1.row_dimensions[row].height = 30
    row += 1

row += 1
section_header(ws1, row, 1, "SOCIAL PROOF & EARNED MEDIA", 3); row += 1
fields_sp = [
    ("KOLs noi bat",       "Chi Pu, Van Mai Huong, Bich Phuong, Chau Bui, Linh An, Bang Di, Co Em Trendy, Nguyen Lam Thao Tam,..."),
    ("Bao chi / Editorial","L'Officiel Vietnam, ELLE Vietnam, Marie Claire Vietnam, Miss Universe Vietnam"),
    ("Su kien lon",         "WeChoice, Van Xuan, Le Hoi Ao Dai, Marie Claire Fashion Night Out"),
    ("Wedding / Bridal",    "The Planners, TIE Men; Co Em Trendy bridal post ~9K reactions"),
    ("Diem manh hien tai",  "Ao dai & occasion wear cao cap; swimwear dang noi len; multi-tier KOL reach"),
    ("Diem can cai thien",  "Low UGC, chua co khach quoc te, it Reels/video, followers thap so voi do phu editorial"),
]
for i, (lbl, val) in enumerate(fields_sp):
    label_cell(ws1, row, 1, lbl)
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws1, row, 2, val, row_shade=(i % 2 == 0))
    ws1.row_dimensions[row].height = 30
    row += 1

# ── Sheet 2: Cac Ben Lien Quan ──────────────────────────────────────────────
ws2 = wb.create_sheet("Stakeholders")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 26
ws2.column_dimensions["D"].width = 46

ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value     = "MAISON DENUDE - CAC BEN LIEN QUAN (STAKEHOLDERS)"
c.font      = Font(name=FONT_MAIN, size=14, bold=True, color=GOLD)
c.fill      = fill(DARK_BG)
c.alignment = center()
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:D2")
c = ws2["A2"]
c.value     = "Nguon: WhatsApp project chat (5/16-6/18) + scope.md"
c.font      = Font(name=FONT_MAIN, size=9, italic=True, color="FF888888")
c.fill      = fill(MID_BG)
c.alignment = center()

row = 4
headers = ["Nguoi", "Vai tro", "Quyen quyet dinh", "Ghi chu"]
for ci, h in enumerate(headers, 1):
    c = ws2.cell(row=row, column=ci, value=h)
    c.font      = hdr_font(10, True, WHITE)
    c.fill      = fill(GOLD)
    c.alignment = center(wrap=False)
    c.border    = thin_border()
ws2.row_dimensions[row].height = 22
row += 1

stakeholders = [
    ("Chi Chi (Chi Bui)", "Founder / Giam doc, Fashion Designer",
     "CAO NHAT - phe duyet cuoi cung",
     "Quyet dinh chien luoc, thuong hieu, keyword. Phong cach: fast-moving, decisive, chủ động debate ky thuat (RedNote, SEMrush). Khong han che - giai thich WHY, khong chi KET QUA."),
    ("Chi Michelle (Miseo)", "Marketing / Dai dien phia doi tac",
     "Phoi hop + phe duyet",
     "Email: mitodaseo@gmail.com. Dieu phoi lich, chuyen file, nhac Chi duyet. Hiem khi khoi xuong chien luoc."),
    ("Phan Ha (be Ha)", "In-house SEO / Content Writer (phia Maison)",
     "Thuc thi noi dung",
     "Viet blog content thuong hieu. Uu tien giong artisanal/editorial; muon bai nghien cuu sau ve loai trang phuc."),
    ("Nguyen Thai Thieu", "Dev + SEO Lead (Ben B)",
     "Giao hang ky thuat",
     "Build website, cai tracking, publish blog, cau truc keyword."),
    ("Matthew", "Dev support (Ben B)", "Ho tro ky thuat", ""),
    ("Maddy", "Research (Ben B)", "Nghien cuu thi truong",
     "Digital Marketing Playbook - 3 thi truong xuat khau."),
]
for i, (name, role, auth, note) in enumerate(stakeholders):
    shade = GREY_ROW if i % 2 == 0 else WHITE
    for ci, val in enumerate([name, role, auth, note], 1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.font      = body_font(10, bold=(ci == 1))
        c.fill      = fill(shade)
        c.alignment = left(wrap=True)
        c.border    = thin_border()
    ws2.row_dimensions[row].height = 52
    row += 1

row += 1
section_header(ws2, row, 1, "QUY TAC LAM VIEC", 4); row += 1
rules = [
    "Van de chien luoc / cuoi cung: gui cho Chi Chi",
    "Lich hop, duyet, chuyen file: gui cho Chi Michelle",
    "Noi dung blog (du thao, brief): phoi hop voi Phan Ha",
    "Quy tac 48h: Maison phai phan hoi deliverable trong 48 gio lam viec; im lang = tu dong chap nhan",
    "Hop weekly: Thu 7, 15:00 (Maison can tham du)",
]
for i, rule in enumerate(rules):
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws2.cell(row=row, column=1, value="  * " + rule)
    c.font      = body_font(10, color=BLACK)
    c.fill      = fill(GOLD_LIGHT if i % 2 == 0 else WHITE)
    c.alignment = left(wrap=True)
    c.border    = thin_border()
    ws2.row_dimensions[row].height = 20
    row += 1

# ── Sheet 3: Hop Dong & Pham Vi ────────────────────────────────────────────
ws3 = wb.create_sheet("Hop Dong Pham Vi")
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 52
ws3.column_dimensions["C"].width = 14

ws3.merge_cells("A1:C1")
c = ws3["A1"]
c.value     = "MAISON DENUDE - HOP DONG & PHAM VI DU AN (PHASE 1)"
c.font      = Font(name=FONT_MAIN, size=14, bold=True, color=GOLD)
c.fill      = fill(DARK_BG)
c.alignment = center()
ws3.row_dimensions[1].height = 32

row = 3
section_header(ws3, row, 1, "THONG TIN HOP DONG", 3); row += 1
contract_info = [
    ("So hop dong",        "01/2026/HDCTV-MD"),
    ("Ben A (khach hang)", "Maison Denude"),
    ("Ben B (nha thau)",   "Nguyen Thai Thieu / Matthew"),
    ("Gia tri Phase 1",    "8.000.000 VND (Net)"),
    ("Dot thanh toan 1",   "2.000.000 VND - khi ky hop dong (coc)"),
    ("Dot thanh toan 2",   "7.500.000 VND - khi nghiem thu cuoi tuan 7"),
    ("Thoi han",           "7 tuan (kickoff 2026-05-31, go-live muc tieu ~2026-07-19)"),
    ("Ten mien",           "KHONG bao gom - Maison tu mua (OQ-002)"),
    ("Add-on A",           "Google Calendar booking integration: +500.000 VND (tuy chon)"),
    ("Add-on B",           "Tinh nang Wishlist (server-stored, an danh): +1.000.000 VND (tuy chon)"),
    ("Bao tri hang nam",   "Mien phi 12 thang dau; sau do 1.000.000 VND/nam"),
]
for i, (lbl, val) in enumerate(contract_info):
    label_cell(ws3, row, 1, lbl)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws3, row, 2, val, row_shade=(i % 2 == 0))
    ws3.row_dimensions[row].height = 22
    row += 1

row += 1
section_header(ws3, row, 1, "PHAM VI TRONG HOP DONG (IN SCOPE)", 3); row += 1
in_scope = [
    ("SEO & Tracking",     "GA4, GTM, GSC, Google Business Profile; backlink building; keyword map (7 tu khoa); blog publish & structure"),
    ("Website Build",      "Astro hybrid -> Cloudflare Pages; trang: /, /blog, /collections, /booking, /admin (CMS collections)"),
    ("He thong dat lich",  "Booking form + email notification (Supabase Edge Function); popup tu dong (~30 giay)"),
    ("Media & Storage",    "Cloudflare R2 (anh/video) + Supabase (URL + metadata)"),
    ("Playbook",           "Digital Marketing Playbook cho 3 thi truong xuat khau (tieng Viet)"),
    ("Handover docs",      "Huong dan them blog (MDX), xem dashboard bookings, bao cao keyword baseline (tieng Viet)"),
]
for i, (lbl, val) in enumerate(in_scope):
    label_cell(ws3, row, 1, lbl)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws3, row, 2, val, row_shade=(i % 2 == 0))
    ws3.row_dimensions[row].height = 30
    row += 1

row += 1
section_header(ws3, row, 1, "NGOAI PHAM VI (OUT OF SCOPE)", 3); row += 1
out_scope = [
    "Mua ten mien / custom email domain",
    "Viet noi dung blog (Maison tu chuan bi; Thieu cau truc & publish)",
    "CMS tu chinh sua blog/page (chi CMS collections)",
    "Da ngon ngu ZH/KO (hoan lai - OQ-005)",
    "Quan ly mang xa hoi / quang cao / KOL outreach",
    "Dashboard metrics rieng cho team (metrics giao qua bao cao Thieu chuan bi)",
]
for i, item in enumerate(out_scope):
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws3.cell(row=row, column=1, value="  [X]  " + item)
    c.font      = body_font(10, color="FF993333")
    c.fill      = fill("FFFFF5F5" if i % 2 == 0 else WHITE)
    c.alignment = left(wrap=True)
    c.border    = thin_border()
    ws3.row_dimensions[row].height = 20
    row += 1

row += 1
section_header(ws3, row, 1, "KPI NGHIEM THU", 3); row += 1
kpis = [
    ("Tracking",     "4/4 cong cu live: GTM, GA4, GSC, Google Business Profile"),
    ("Website",      "Tat ca trang live tren Cloudflare Pages: /, /blog, /collections, /booking, /admin"),
    ("Booking",      "Form -> email functional (test submission confirmed)"),
    ("Noi dung SEO", "Toan bo blog content cua Maison duoc publish + cau truc dung keyword muc tieu"),
    ("Keyword",      "Bao cao baseline rankings cho du 7 keywords"),
    ("Performance",  "Lighthouse > 90 (mobile)"),
    ("Post-launch",  "Top 10 cho >= 3 target keywords sau 3 thang go-live"),
]
for i, (lbl, val) in enumerate(kpis):
    label_cell(ws3, row, 1, lbl)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    value_cell(ws3, row, 2, val, row_shade=(i % 2 == 0))
    ws3.row_dimensions[row].height = 22
    row += 1

ws1.freeze_panes = "A4"
ws2.freeze_panes = "A5"
ws3.freeze_panes = "A4"

ws1.sheet_properties.tabColor = "C9A84C"
ws2.sheet_properties.tabColor = "4C72C9"
ws3.sheet_properties.tabColor = "5CA84C"

out_path = r"C:\maison\Maison_Denude_Ho_So_Doanh_Nghiep.xlsx"
wb.save(out_path)
print("saved:", out_path)
