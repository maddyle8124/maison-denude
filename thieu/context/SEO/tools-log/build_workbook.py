# -*- coding: utf-8 -*-
"""
Tao workbook 'Keyword thi truong' cho Maison Denude.
2 sheet: (1) Keyword thi truong tieng Anh, (2) De xuat keyword tieng Viet.
Du lieu: SimilarWeb keywords-overview (volume + clicks, T3-T5/2026, WW & VN)
         + Google Trends 12 thang qua SerpApi (chi so xu huong 0-100, WW & VN).
Tat ca so lieu deu co log goc trong thu muc data_logs/.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "MaisonDenude_Keyword_ThiTruong_2026-06.xlsx")

# ----- Styles -----
INK = "1F2A33"        # dark slate
HEAD_FILL = PatternFill("solid", fgColor=INK)
TITLE_FONT = Font(name="Calibri", size=15, bold=True, color=INK)
SUB_FONT = Font(name="Calibri", size=10, color="55606A", italic=True)
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=10, color="222222")
KW_FONT = Font(name="Calibri", size=10, bold=True, color="111111")
thin = Side(style="thin", color="D5DADF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

HEADERS = [
    ("STT", 5),
    ("Tu khoa", 20),
    ("Nhom chu de", 18),
    ("Volume toan cau\n(/thang)", 14),
    ("Volume Viet Nam\n(/thang)", 15),
    ("Luong click toan cau\n(/thang)", 16),
    ("Xu huong 12 thang\nToan cau (0-100)", 15),
    ("Xu huong 12 thang\nViet Nam (0-100)", 15),
    ("Search intent", 13),
    ("Do kho\n(0-100)", 9),
    ("CPC (USD)", 13),
    ("Nguon du lieu", 26),
    ("Ghi chu", 40),
]
# Vietnamese display labels for headers (overrides above ascii safe names)
HEADERS_VI = [
    "STT",
    "Từ khoá",
    "Nhóm chủ đề",
    "Volume toàn cầu\n(/tháng)",
    "Volume Việt Nam\n(/tháng)",
    "Lượng click toàn cầu\n(/tháng)",
    "Xu hướng 12 tháng\nToàn cầu (0–100)",
    "Xu hướng 12 tháng\nViệt Nam (0–100)",
    "Search intent",
    "Độ khó\n(0–100)",
    "CPC (USD)",
    "Nguồn dữ liệu",
    "Ghi chú",
]
WIDTHS = [5, 20, 18, 14, 15, 16, 15, 15, 13, 9, 14, 26, 42]

SRC = "SimilarWeb keywords-overview (T3–T5/2026) + Google Trends 12 tháng (SerpApi)"

# Each row: STT, keyword, group, vol_ww, vol_vn, clicks_ww, trend_ww, trend_vn,
#           intent, difficulty, cpc, source, note
# vol/clicks: int hoac chuoi "Không đo được". trend: int hoac "—".

EN_ROWS = [
    [1, "cocktail dress", "Evening & Occasion", 149265, 832, 49566, 16, 1, "Giao dịch", 24, "$0.01 – $2.85", SRC,
     "Cầu rất lớn toàn cầu, intent mua rõ. Đỉnh nhẹ mùa lễ/Tết (T2/2026). Volume VN nhỏ."],
    [2, "haute couture", "Couture", 95734, 2763, 15389, 5, 2, "Thông tin", 40, "$0.01 – $16.45", SRC,
     "Phần lớn là intent tìm hiểu (couture là gì), không phải mua. Độ khó cao. Dùng cho bài định vị."],
    [3, "formal dress", "Evening & Occasion", 75716, 188, 24215, 33, 2, "Giao dịch", 20, "$0.02 – $2.41", SRC,
     "Cầu lớn, intent mua. Xu hướng tăng mạnh quanh mùa tiệc (T2–T4)."],
    [4, "ao dai", "Áo dài", 68647, 8781, 15750, 3, "—", "Thông tin", 16, "—", SRC,
     "Head term: ~63–72K/tháng toàn cầu, gần 100% intent thông tin. VN: ~8.781/tháng. Bắt traffic bằng bài guide."],
    [5, "evening dresses", "Evening & Occasion", 22188, "Không đo được", 10354, 8, 1, "Giao dịch", 17, "$0.02 – $2.61", SRC,
     "Biến thể số nhiều của 'evening dress', intent mua. VN: không đo được (quá nhỏ)."],
    [6, "evening gown", "Evening & Occasion", 19881, 290, 5281, 6, 1, "Giao dịch", 33, "$0.02 – $3.52", SRC,
     "Cầu khá, intent mua. Độ khó trung bình–cao."],
    [7, "evening dress", "Evening & Occasion", 14874, "Không đo được", 5342, 15, 4, "Giao dịch", 20, "$0.03 – $2.61", SRC,
     "Từ lõi nhóm dạ hội, intent mua. Đỉnh mạnh mùa Tết. VN: không đo được."],
    [8, "bespoke suit", "Bespoke tailoring", 11953, 440, 1597, 2, "—", "Địa phương", 46, "$0.04 – $11.10", SRC,
     "Định vị đặt may cao cấp. Intent thiên về địa phương/thương mại. Độ khó cao (đối thủ Savile Row…)."],
    [9, "evening wear", "Evening & Occasion", 4499, "Không đo được", 446, 3, "—", "Giao dịch", 23, "$0.02 – $4.61", SRC,
     "Cụm occasion-wear, intent mua/địa phương. Cầu vừa, cạnh tranh vừa."],
    [10, "bespoke tailor", "Bespoke tailoring", 3835, "Không đo được", 214, 1, "—", "Địa phương", 41, "$0.07 – $7.74", SRC,
     "Đặt may riêng, intent địa phương. Khớp định vị xưởng may Saigon."],
]

VN_ROWS = [
    [1, "áo dài", "Áo dài", 43521, 37819, 16679, 3, 43, "Thông tin", 25, "$0.02 – $0.13", SRC,
     "Head term tiếng Việt, đỉnh Tết (T1–T2). Gần 100% intent thông tin. Cạnh tranh cao nội địa."],
    [2, "áo dài cách tân", "Áo dài", 9921, 9754, 2976, "—", 1, "Thông tin", 19, "$0.02 – $0.10", SRC,
     "Chủ đề áo dài hiện đại, cầu lớn (đỉnh T3/2026). Hợp định vị 'đương đại' của Maison."],
    [3, "áo dài nam", "Áo dài", 6010, 5711, 2486, "—", 5, "Thông tin", "—", "$0.02 – $0.26", SRC,
     "Áo dài nam/chú rể có cầu thật. Bổ trợ cho gói cưới."],
    [4, "váy dạ hội", "Evening & Occasion", 4963, 4691, 2041, "—", 0, "Giao dịch", 56, "$0.04 – $0.10", SRC,
     "Đồng nghĩa 'đầm dạ hội', intent mua. Độ khó cao hơn."],
    [5, "áo dài truyền thống", "Áo dài", 3015, 2959, 1139, "—", 1, "Thông tin", 3, "$0.02 – $0.07", SRC,
     "Cụm di sản, độ khó thấp → dễ lên top bằng bài nội dung."],
    [6, "đầm dạ hội", "Evening & Occasion", 1863, 1784, 959, "—", 0, "Giao dịch", "—", "$0.03 – $0.11", SRC,
     "Cụm dạ hội phổ biến, intent mua. Hợp dòng occasion wear."],
    [7, "áo dài đẹp", "Áo dài", 1673, 1673, 1114, "—", 3, "Thông tin", 20, "$0.02 – $0.09", SRC,
     "Cụm cảm hứng/tham khảo mẫu. Tốt cho bài lookbook."],
    [8, "áo dài cưới", "Bridal", 1237, 1184, 762, "—", 1, "Thông tin / Giao dịch", 5, "$0.06 – $0.12", SRC,
     "Áo dài cưới, độ khó thấp, intent chuyển đổi tốt. Ưu tiên cho nhóm cưới."],
    [9, "đầm dự tiệc", "Evening & Occasion", 590, 590, 271, "—", 1, "Giao dịch", 4, "$0.01 – $0.12", SRC,
     "Đầm dự tiệc, intent mua, độ khó thấp → long-tail dễ rank."],
    [10, "thuê áo dài", "Áo dài", 584, 467, 295, "—", 2, "Địa phương", 2, "$0.07 – $0.15", SRC,
     "Intent thuê (địa phương). Cân nhắc nếu Maison có dịch vụ cho thuê; nếu không, dùng để hiểu thị trường."],
]

INTRO_EN = [
    "Maison Denude — Keyword thị trường (tiếng Anh), đã xác minh số liệu",
    "Bộ từ khoá đại trà của ngành (váy/đầm dạ hội, couture, áo dài, đặt may), tổng hợp từ các topic Google Trends "
    "rồi rút gọn còn 10 từ khoá có volume/traffic cao nhất sau khi xác minh bằng SimilarWeb.",
]
INTRO_VN = [
    "Maison Denude — Đề xuất keyword tiếng Việt, đã xác minh số liệu",
    "Bộ từ khoá tiếng Việt của thị trường nội địa, cùng phương pháp với bảng tiếng Anh: tổng hợp từ Google Trends "
    "Việt Nam rồi rút gọn còn 10 từ khoá có volume/traffic cao nhất theo SimilarWeb (chế độ Việt Nam).",
]

METHOD = [
    "Nguồn & phương pháp:",
    "• Volume & lượng click (traffic): SimilarWeb keywords-overview, dữ liệu tháng mới nhất (T3–T5/2026), hai chế độ Toàn cầu (ww) và Việt Nam (vn). Volume = lượt tìm/tháng (trung bình 3 tháng). 'Không đo được' = API trả rỗng (quá nhỏ để đo).",
    "• Xu hướng 12 tháng: Google Trends qua SerpApi (TIMESERIES, 'today 12-m'), hai chế độ Toàn cầu và Việt Nam. Đây là CHỈ SỐ TƯƠNG ĐỐI 0–100, KHÔNG phải lượt tuyệt đối. Bảng tiếng Việt lấy 'áo dài' làm mốc (=43) nên các từ nhỏ hơn đọc thấp — cầu thật xem ở cột Volume.",
    "• Search intent & Độ khó & CPC: từ SimilarWeb. CPC quy đổi ra USD. Độ khó 0–100 (cao = khó lên top).",
    "• Toàn bộ dữ liệu thô lưu tại thư mục data_logs/ để đối chiếu.",
]


def build_sheet(wb, title, intro, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ncol = len(HEADERS_VI)
    last_col = get_column_letter(ncol)

    # column widths
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws.cell(r, 1, intro[0]); c.font = TITLE_FONT; c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws.cell(r, 1, intro[1]); c.font = SUB_FONT; c.alignment = WRAP
    ws.row_dimensions[r].height = 40
    r += 1

    # method block
    for line in METHOD:
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws.cell(r, 1, line)
        c.font = Font(name="Calibri", size=9, color="55606A",
                      bold=line.endswith(":"))
        c.alignment = WRAP
        ws.row_dimensions[r].height = 30 if len(line) > 90 else 15
        r += 1
    r += 1  # spacer

    # header
    head_row = r
    for j, label in enumerate(HEADERS_VI, start=1):
        c = ws.cell(head_row, j, label)
        c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[head_row].height = 30
    r += 1

    # data
    for row in rows:
        for j, val in enumerate(row, start=1):
            c = ws.cell(r, j, val)
            c.border = BORDER
            if j == 1:
                c.alignment = CENTER; c.font = CELL_FONT
            elif j == 2:
                c.alignment = Alignment(vertical="center"); c.font = KW_FONT
            elif j in (4, 5, 6):      # volume / clicks
                c.font = CELL_FONT
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"; c.alignment = RIGHT
                else:
                    c.alignment = CENTER
            elif j in (7, 8, 10):     # trend / difficulty
                c.font = CELL_FONT; c.alignment = CENTER
            elif j in (3, 9, 11):     # group / intent / cpc
                c.font = CELL_FONT; c.alignment = CENTER
            else:                      # source / note
                c.font = CELL_FONT; c.alignment = WRAP
        ws.row_dimensions[r].height = 46
        r += 1

    ws.freeze_panes = ws.cell(head_row + 1, 1)
    ws.sheet_view.showGridLines = False
    return ws


def main():
    wb = Workbook()
    build_sheet(wb, "Keyword thị trường", INTRO_EN, EN_ROWS, first=True)
    build_sheet(wb, "Đề xuất keyword tiếng Việt", INTRO_VN, VN_ROWS)
    wb.save(OUT)
    print("Saved:", OUT)


if __name__ == "__main__":
    main()
