# -*- coding: utf-8 -*-
"""Build Maison Denude SEO Research & Strategy workbook (client-facing, VN prose)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Maison_Denude_SEO_Research_Strategy.xlsx")

FONT = "Arial"
# palette
NAVY = "1F3A5F"      # header bg
GOLD = "C9A24B"      # accent / section
LIGHT = "F2EFE9"     # band
INSIGHT = "EAF1F8"   # insight block
EVID = "FBF7EE"      # evidence block
DEC = "EDF5EE"       # decision block
WHITE = "FFFFFF"
GREY = "6B6B6B"

thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def fill(hexc):
    return PatternFill("solid", fgColor=hexc)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
TOP = Alignment(vertical="top")
CENTER = Alignment(vertical="center", horizontal="center")

wb = Workbook()
wb.remove(wb.active)


def new_sheet(title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    return ws


def title_block(ws, title, subtitle=None, span=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.font = F(16, True, NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    r = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(2, 1, subtitle)
        s.font = F(10, False, GREY, italic=True)
        s.alignment = WRAP
        ws.row_dimensions[2].height = 30
        r = 3
    return r + 1  # leave a blank row


def table(ws, start_row, headers, rows, widths=None, wrap_cols=None,
          header_fill=NAVY, header_color=WHITE):
    wrap_cols = wrap_cols or set()
    # header
    for j, h in enumerate(headers, 1):
        c = ws.cell(start_row, j, h)
        c.font = F(10, True, header_color)
        c.fill = fill(header_fill)
        c.alignment = WRAP_C
        c.border = BORDER
    ws.row_dimensions[start_row].height = 28
    # body
    for i, row in enumerate(rows, 1):
        rr = start_row + i
        band = LIGHT if i % 2 == 0 else WHITE
        for j, val in enumerate(row, 1):
            c = ws.cell(rr, j, val)
            c.font = F(10)
            c.fill = fill(band)
            c.alignment = WRAP if (j in wrap_cols) else TOP
            c.border = BORDER
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return start_row + len(rows) + 1


def section_label(ws, row, text, span=6, color=GOLD):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = F(12, True, WHITE)
    c.fill = fill(color)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def note_line(ws, row, text, span=6, color=GREY, italic=True, height=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = F(9, False, color, italic=italic)
    c.alignment = WRAP
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


# =====================================================================
# 0. MỤC LỤC
# =====================================================================
ws = new_sheet("0. Mục lục")
r = title_block(ws, "Maison Denude — Nghiên Cứu & Chiến Lược SEO",
    "Tài liệu tổng hợp (gộp từ Bản đồ SEO Landscape + 10 Blog Brief) — phiên bản trình bày cho khách hàng. "
    "Nguồn dữ liệu: SimilarWeb MCP + SerpApi/Google Trends thực tế + SERP đo thật. Ngày thực hiện: 11/6/2026.", span=3)

r = section_label(ws, r, "Một dòng chiến lược", span=3)
r = note_line(ws, r,
    "KHÔNG đua từ khoá lớn (head term). Chiếm nhóm từ khoá dài, cụ thể, có ý định mua cao + định vị "
    "“bespoke áo dài Saigon” cho khách quốc tế/expat — nơi gần như không có nhà may cao cấp nào cạnh tranh — "
    "bằng mô hình cụm nội dung (hub-and-spoke) để đạt KPI Top-10 cho ≥3 từ khoá trong 3 tháng.",
    span=3, color="000000", italic=False, height=46)
r += 1

r = section_label(ws, r, "Danh mục các sheet", span=3)
idx_rows = [
    ["1. Tổng quan & Quyết định", "Bản tóm tắt cho người bận: insight → bằng chứng (số thật) → quyết định. ĐỌC SHEET NÀY TRƯỚC."],
    ["2. Audit 34 từ khoá", "Đánh giá toàn bộ 34 từ khoá khách cung cấp, đối chiếu volume thực tế."],
    ["3. Cạnh tranh SERP", "Ai đang đứng trang 1 Google (đo theo VN·US·AU) — và khoảng trống cho Maison theo thị trường."],
    ["4. Shortlist ưu tiên (Top 10)", "10 từ khoá ưu tiên (đầu vào cho 10 bài blog), xếp theo cơ hội × giá trị."],
    ["5. Trends · Geo · Related", "Mức quan tâm thực 12 tháng, demand đến từ đâu (Úc/Mỹ), từ khoá liên quan."],
    ["6. Bản đồ khoảng trống", "Các khoảng trống thị trường và cách Maison khai thác."],
    ["7. Câu hỏi khách hỏi (PAA)", "Ngân hàng câu hỏi thật khách quốc tế hay tìm + đáp án — nguyên liệu viết bài."],
    ["8. 10 Blog Brief", "Định hướng chi tiết 10 bài viết (mỗi bài 1 dòng)."],
    ["9. Kiến trúc cụm chủ đề", "Vì sao thiết kế 3 trụ + 7 bài hỗ trợ, và quy tắc liên kết."],
    ["10. Lịch xuất bản", "Thứ tự đăng bài đề xuất để rank nhanh."],
    ["11. Checklist SEO kỹ thuật", "Danh sách kiểm tra khi đăng mỗi bài (dành cho đội kỹ thuật)."],
    ["12. Phương pháp & Độ tin cậy", "Cách thu thập & đọc số liệu; mức độ tin cậy."],
    ["13. Bàn giao & Ghi chú", "Liên kết công việc với Maddy, thay thế bảng cũ, cập nhật sau go-live."],
    ["14. Thuật ngữ (Glossary)", "Giải nghĩa các thuật ngữ SEO bằng tiếng Việt dễ hiểu."],
]
r = table(ws, r, ["Sheet", "Nội dung"], idx_rows, widths=[34, 78], wrap_cols={2})
ws.column_dimensions["C"].width = 4


# =====================================================================
# 1. TỔNG QUAN & QUYẾT ĐỊNH  (Master)
# =====================================================================
ws = new_sheet("1. Tổng quan & Quyết định")
W = [22, 64, 26]  # label | content | ref
for j, wdt in enumerate(W, 1):
    ws.column_dimensions[get_column_letter(j)].width = wdt
r = title_block(ws, "Tổng Quan & Quyết Định Chiến Lược",
    "Cấu trúc mỗi mục: PHÁT HIỆN (nói bằng ngôn ngữ kinh doanh) → BẰNG CHỨNG (số liệu thật) → QUYẾT ĐỊNH (ta làm gì). "
    "Số liệu chi tiết nằm ở các sheet được trỏ tới ở cột bên phải.", span=3)

# core question
r = section_label(ws, r, "Câu hỏi cốt lõi & Mục tiêu (KPI)", span=3)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1, "Maison Denude có thể lên top Google nhanh ở đâu để đạt KPI: Top-10 cho ≥3 từ khoá trong 3 tháng sau go-live?")
c.font = F(11, True, NAVY); c.alignment = WRAP; ws.row_dimensions[r].height = 30
r += 2

def insight_block(ws, r, title, insight, evidence, decision, ref):
    # title bar
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, title); c.font = F(11, True, WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 22
    r += 1
    for label, text, bg in [("💡 Phát hiện", insight, INSIGHT),
                            ("📊 Bằng chứng", evidence, EVID),
                            ("✅ Quyết định", decision, DEC)]:
        lc = ws.cell(r, 1, label); lc.font = F(10, True); lc.fill = fill(bg)
        lc.alignment = WRAP; lc.border = BORDER
        tc = ws.cell(r, 2, text); tc.font = F(10); tc.fill = fill(bg)
        tc.alignment = WRAP; tc.border = BORDER
        rc = ws.cell(r, 3, ref if label.startswith("📊") else "")
        rc.font = F(9, False, GREY, italic=True); rc.fill = fill(bg)
        rc.alignment = WRAP; rc.border = BORDER
        ws.row_dimensions[r].height = max(30, 14 * (len(text) // 60 + 1))
        r += 1
    return r + 1

r = section_label(ws, r, "3 phát hiện quyết định chiến lược", span=3)
r = insight_block(ws, r,
    "1. Người ta tìm “áo dài” để HIỂU, không phải để MUA",
    "Từ khoá lớn nhất “ao dai” có 62.244–72.296 lượt tìm/tháng toàn cầu, nhưng 100% là tìm để tìm hiểu (không phải mua).",
    "SimilarWeb T3–T5/2026: 72.296 (T3) · 62.244 (T4) · 71.402 (T5); informational intent = 100% mỗi tháng, giao dịch = 0.",
    "Không đua head term. Bắt traffic này bằng bài guide (vd bài ‘ý nghĩa áo dài’) rồi điều hướng xuống trang đặt lịch.",
    "→ Sheet ‘2. Audit 34 từ khoá’")
r = insight_block(ws, r,
    "2. Các từ định vị cao cấp gần như không có người tìm trực tiếp → cạnh tranh thấp",
    "Những từ gắn “bespoke / saigon / couture / occasion” có mức quan tâm rất nhỏ. Với ngách long-tail, mức thấp = CƠ HỘI lên top nhanh, không phải tin xấu.",
    "Google Trends 12 tháng (thang chuẩn hoá, wedding ao dai=50): bespoke ao dai=0 · tailored ao dai=0 · vietnamese couture=1 · evening wear saigon=1 · ao dai saigon=6. SimilarWeb trả rỗng cho bespoke ao dai & ao dai saigon.",
    "Phủ nhiều bài long-tail cụ thể (không đua 1–2 từ khoá lớn) để lên top nhanh từng bài.",
    "→ Sheet ‘2. Audit 34 từ khoá’, ‘5. Trends · Geo · Related’")
r = insight_block(ws, r,
    "3. Khoảng trống SERP thay đổi theo thị trường",
    "Đo thật trang 1 Google theo ma trận VN·US·AU (11/6/2026): kết quả khác nhau rõ giữa thị trường nội địa và diaspora.",
    "Tại VN, atelier Việt đã đứng top (Local Pack + Maydo/Duan; Nicole Bridal rank cả product lẫn blog) → cần cạnh tranh + Google Business Profile. Tại US/AU, top do seller diaspora + Amazon chiếm, không có atelier Saigon.",
    "Khoảng trống đúng của Maison: góc “bespoke made in Saigon, thử trực tiếp” ở thị trường diaspora; còn ở VN thắng bằng nội dung sâu hơn + GBP.",
    "→ Sheet ‘3. Cạnh tranh SERP’")

# audience
r = section_label(ws, r, "Đối tượng SEO nhắm tới", span=3)
for label, text in [
    ("Nhóm 1 — Expat tại VN",
     "Người nước ngoài sống ở TP.HCM/Hà Nội, muốn đồ bespoke chuẩn quốc tế. Giai đoạn cuối phễu (sẵn sàng đặt). VD tìm: “best ao dai tailor in saigon”."),
    ("Nhóm 2 — Khách quốc tế đến VN",
     "Du khách chi tiêu cao, tìm món unique; có người đặt trước khi bay sang, có người đến nơi mới tìm. Đây là “pre-trip booking funnel” (đặt may trước chuyến đi). VD: “custom ao dai saigon”, “how much does an ao dai cost”."),
    ("Hệ quả nội dung",
     "Viết tiếng Anh, kiểu editorial dễ đọc; luôn có tín hiệu địa lý Saigon/Ho Chi Minh City/District 1 + nút đặt lịch (CTA). Định dạng thắng: guide dài có cấu trúc + ảnh đẹp (Maison có sẵn ảnh KOL/editorial)."),
]:
    lc = ws.cell(r, 1, label); lc.font = F(10, True); lc.alignment = WRAP; lc.fill = fill(LIGHT); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    tc = ws.cell(r, 2, text); tc.font = F(10); tc.alignment = WRAP; tc.border = BORDER
    ws.row_dimensions[r].height = max(30, 13 * (len(text)//70 + 2))
    r += 1
r += 1

# conclusion
r = section_label(ws, r, "Kết luận chiến lược", span=3)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1,
    "Maison KHÔNG đua từ khoá lớn. Maison chiếm “từ khoá dài có ý định mua cao + định vị địa lý Saigon + đối tượng "
    "khách quốc tế/expat”. Mô hình nội dung: cụm chủ đề gồm 3 bài trụ (pillar) + 7 bài hỗ trợ long-tail liên kết "
    "nội bộ, tạo mạng lưới đủ mạnh để Google hiểu Maison là chuyên gia về bespoke áo dài tại Saigon, bán ra toàn cầu.")
c.font = F(11, True, NAVY); c.fill = fill(DEC); c.alignment = WRAP; c.border = BORDER
ws.row_dimensions[r].height = 70
r += 2

# corrections
r = section_label(ws, r, "3 hiệu chỉnh quan trọng sau khi đối chiếu số liệu thật", span=3)
corr = [
    ["1. “traditional ao dai” mạnh bất ngờ",
     "Điểm Trends 57/50 — CAO HƠN cả “wedding ao dai”. → Nâng ưu tiên: làm bài “ý nghĩa áo dài” làm nam châm thu hút (bài #10)."],
    ["2. “vietnamese couture” bị thổi phồng",
     "Brief khách ghi “↑33” nhưng thật ra chỉ là 1 tuần đột biến sau Couture Week; Trends 12 tháng thật = 1/50. → KHÔNG coi là từ kéo traffic; chỉ dùng làm góc định vị thương hiệu trong Pillar C."],
    ["3. “wedding ao dai” là từ mạnh nhất + cầu từ Úc & Mỹ",
     "Trends 50/50 (mạnh nhất toàn cụm). Demand đến gần như 100% từ Úc & Mỹ (diaspora Việt), Việt Nam <1. → Pillar B dùng “wedding ao dai” làm từ chính + nhắm cô dâu gốc Việt ở Úc/Mỹ; bổ sung tín hiệu áo dài chú rể (ao dai nam=43)."],
]
for label, text in corr:
    lc = ws.cell(r, 1, label); lc.font = F(10, True); lc.alignment = WRAP; lc.fill = fill(EVID); lc.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    tc = ws.cell(r, 2, text); tc.font = F(10); tc.alignment = WRAP; tc.border = BORDER
    ws.row_dimensions[r].height = max(30, 13 * (len(text)//70 + 2))
    r += 1
ws.freeze_panes = "A4"


# =====================================================================
# 2. AUDIT 34 TỪ KHOÁ
# =====================================================================
ws = new_sheet("2. Audit 34 từ khoá")
r = title_block(ws, "Audit 34 Từ Khoá (đối chiếu volume thực tế)",
    "Mọi số đều có nguồn (data_logs/). SW vol = volume SimilarWeb (toàn cầu, EN, T3–T5/2026); “rỗng” = quá nhỏ để đo. "
    "Trends 12m = điểm Google Trends 12 tháng đã chuẩn hoá (wedding ao dai=50 làm neo). SERP US = vị trí trang 1 đo thật 11/6/2026. "
    "Lưu ý: điểm 0 nghĩa là “rất nhỏ so với neo”, KHÔNG phải “không ai tìm” — với long-tail, điểm thấp = cạnh tranh thấp = cơ hội rank nhanh.", span=8)
hdr = ["#", "Từ khoá", "SW vol", "Trends 12m", "SERP US", "Intent", "Verdict", "Ghi chú (có số)"]
kw = [
    [1, "bespoke ao dai", "rỗng (data[]=[])", "0/50", "—", "Trans/Brand", "🏛 Pillar A (head)", "Volume đo được = 0 nhưng là định vị lõi. ‘bespoke ao dai saigon’: ở VN có atelier Việt + Local Pack (cạnh tranh), ở US gần như trống atelier"],
    [2, "ao dai saigon", "rỗng (data[]=[])", "6/50", "—", "Local/Trans", "🏛 Pillar A (geo)", "“Volume cao nhất” trong brief khách là so nội bộ; Trends thật chỉ 6/50 — vẫn đáng vì intent mua + đối thủ atelier ≈ 0"],
    [3, "heritage ao dai", "rỗng", "1/50", "—", "Info", "🔗 Semantic-support", "Điểm 1 — gần như không ai gõ trực tiếp; rải trong pillar"],
    [4, "silk ao dai", "thấp, ngắt quãng", "11/50", "—", "Info/Trans", "⚡ Quick-win", "Trends thật 11 (đỉnh tuần Dec'25=72, Feb'26=50). Có thật nhưng theo mùa → bài “chọn lụa”"],
    [5, "traditional ao dai", "nằm dưới head ‘ao dai’", "57/50", "—", "Info", "⚡ Quick-win (mạnh bất ngờ)", "Điểm 57 — CAO HƠN cả ‘wedding ao dai’! (đỉnh tuần Feb'26=100). Quick-win giá trị nhất nhóm info → ưu tiên bài #10"],
    [6, "ao dai boutique saigon", "rỗng", "0/50", "—", "Local/Trans", "🔗 Semantic-support", "Biến thể #2; gộp vào Pillar A"],
    [7, "tailored ao dai", "rỗng", "0/50", "—", "Trans", "⚡ Quick-win", "Điểm 0 nhưng intent đặt may rõ + đối thủ ≈ 0 → vẫn làm bài long-tail"],
    [8, "designer ao dai saigon", "rỗng", "0/50", "—", "Trans", "🔗 Semantic-support", "Tín hiệu luxury; rải trong Pillar A & C"],
    [9, "bridal ao dai", "thấp", "1/50", "—", "Trans", "🏛→🔗 Semantic-support", "Hiệu chỉnh: Trends thật chỉ 1/50 — cầu thực nằm ở ‘wedding ao dai’ (#12). Hạ từ head xuống phụ"],
    [10, "wedding dress vietnam", "thấp", "~1 (ước)", "—", "Info/Trans", "🔗 Semantic-support", "Rộng; gộp vào Pillar B"],
    [11, "bespoke wedding saigon", "rỗng", "~0 (ước)", "—", "Trans", "⚡ Quick-win", "Đo được = 0 nhưng giao điểm bespoke × bridal × geo = “vàng” + đối thủ ≈ 0"],
    [12, "wedding ao dai", "826(T4)→25(T5); Info 100%", "50/50 (NEO)", "US/AU: seller diaspora; VN: Nicole Bridal", "Info→Trans", "🏛 Pillar B (head)", "Từ khoá mạnh nhất toàn cụm. Trends 12m=50 (đỉnh tuần 100 đầu T4). US/AU không có atelier Saigon → góc ‘made in Saigon’ độc quyền; VN có nhà cưới Việt rank"],
    [13, "the ao dai", "nằm dưới head", "n/a", "—", "Info", "✂️ Drop/gộp", "Quá generic"],
    [14, "bridesmaids ao dai", "rỗng", "0/50", "—", "Trans", "⚡ Quick-win", "Trends 0 nhưng intent rõ + đối thủ ≈ 0 → long-tail conversion"],
    [15, "modern bridal ao dai", "thấp", "0/50", "—", "Info/Trans", "⚡ Quick-win", "Cụm gốc 0; nhưng ‘ao dai modern’ (related query) = 24 → viết theo hướng ‘modern ao dai’"],
    [16, "ao dai wedding", "~ #12", "(gộp #12)", "—", "Info", "🔗 Semantic-support", "Biến thể; related query ‘ao dai wedding’=28"],
    [17, "ao dai bride", "thấp", "0/50", "—", "Info/Trans", "🔗 Semantic-support", "Biến thể; rải trong Pillar B"],
    [18, "silk ao dai bridal", "rỗng", "~0 (ước)", "—", "Trans", "🔗 Semantic-support", "Rải trong bài lụa + bridal"],
    [19, "embroidered dress", "rộng", "n/a (lệch ngách)", "—", "Info", "✂️ Drop/gộp", "Chỉ rải dưới dạng ‘embroidered ao dai’"],
    [20, "ao dai beading", "rỗng", "~0 (ước)", "—", "Info", "🔗 Semantic-support", "Thuật ngữ craft; rải trong bài thủ công"],
    [21, "evening wear saigon", "rỗng", "1/50", "—", "Trans", "🏛 Pillar C (geo head)", "Trends 1 (đỉnh tuần Nov'25=50). Cầu nhỏ nhưng intent đặt may + đối thủ EN ≈ 0 → vẫn làm trụ occasion"],
    [22, "evening dress ho chi minh", "rỗng", "~0 (ước)", "—", "Trans", "🔗 Semantic-support", "Biến thể #21"],
    [23, "artisanal bespoke fashion", "rỗng", "~0 (ước)", "—", "Info/Brand", "🔗 Semantic-support", "Định vị thương hiệu; rải Pillar A & C"],
    [24, "occasion wear vietnam", "rỗng", "1/50", "—", "Trans", "⚡ Quick-win", "(sửa chính tả ‘occasional’→‘occasion’)"],
    [25, "asian inspired dress", "thấp", "n/a (lệch ngách)", "—", "Info", "✂️ Drop/gộp", "Lệch ngách phương Tây; bỏ"],
    [26, "custom evening dress", "thấp", "~0 (ước)", "—", "Trans", "🔗 Semantic-support", "Rải Pillar C"],
    [27, "evening gown vietnam", "rỗng", "~0 (ước)", "—", "Trans", "🔗 Semantic-support", "Rải Pillar C"],
    [28, "contemporary ao dai", "thấp", "1/50", "—", "Info", "⚡ Quick-win", "= #30; related query ‘ao dai modern’=24 cho thấy nên dùng từ ‘modern’"],
    [29, "vietnamese couture", "thấp, blip", "1/50", "—", "Info/Brand", "🅑 Brand (hạ từ Quick-win)", "Hiệu chỉnh: brief khách ghi ‘↑33’ nhưng chỉ là blip 1 tuần (Sep'25=60). Trends thật = 1/50 → KHÔNG phải từ kéo traffic; chỉ làm góc định vị Pillar C"],
    [30, "contemporary ao dai", "thấp", "1/50", "—", "Info", "⚡ Quick-win", "= #28; làm 1 bài"],
    [31, "ao dai haute couture", "rỗng", "~0 (ước)", "—", "Info/Brand", "🔗 Semantic-support", "Rải trong bài couture"],
    [32, "vietnamese haute couture", "rỗng", "~0 (ước)", "—", "Info/Brand", "🔗 Semantic-support", "Rải trong bài couture"],
    [33, "vietnamese dress tailor", "thấp", "~0 (ước)", "—", "Trans", "🔗 Semantic-support", "Rải Pillar A; related query ‘vietnamese dress’=24"],
    [34, "custom ao dai vietnam", "thấp", "0/50", "—", "Trans", "⚡ Quick-win", "Trends 0 nhưng intent đặt may rõ + geo quốc gia → long-tail"],
    ["—", "hand embroidery / hand-guided embroidery / hand embellished / hand beading / beadwork", "rỗng", "~0 (ước)", "—", "Info/Brand", "🔗 Semantic-support", "Bộ thuật ngữ thủ công — rải trong bài craftsmanship để bồi E-E-A-T"],
]
r = table(ws, r, hdr, kw, widths=[5, 30, 22, 13, 30, 12, 22, 60], wrap_cols={2,5,8})
r = note_line(ws, r, "⚠️ Minh bạch dữ liệu: ô ghi “~0 (ước)” là chưa đo Trends riêng — chúng trả rỗng trên SimilarWeb VÀ có cùng pattern near-zero như các từ cùng nhóm đã đo, nên suy ra ~0 thay vì tốn thêm call. 16/34 từ được chấm Trends trực tiếp; phần còn lại suy theo pattern.", span=8, height=44)
r = note_line(ws, r, "Verdict: 🏛 Pillar (bài trụ) · ⚡ Quick-win (bài thắng nhanh) · 🔗 Semantic-support (rải hỗ trợ) · 🅑 Brand (định vị) · ✂️ Drop/gộp (bỏ hoặc gộp).", span=8, height=16)
r = note_line(ws, r, "Tổng kết: 3 Pillar · ~8 Quick-win · còn lại Semantic-support/Brand. 2 hiệu chỉnh lớn: (a) traditional ao dai=57 mạnh nhất nhóm info → nâng ưu tiên; (b) vietnamese couture=1 & bridal ao dai=1 → hạ khỏi vai trò ‘từ kéo traffic’.", span=8, height=30)


# =====================================================================
# 3. CẠNH TRANH SERP
# =====================================================================
ws = new_sheet("3. Cạnh tranh SERP")
r = title_block(ws, "Bản Đồ Cạnh Tranh (ai đang chiếm trang 1 Google)",
    "Đo thật bằng SerpApi (Google engine, 11/6/2026) theo ma trận 3 địa điểm: Việt Nam · US · Úc. Kết quả khác nhau rõ giữa thị trường nội địa (khách đến Saigon) và diaspora (US/AU) — nên đo nhiều nơi. “atelier” = nhà may cao cấp.", span=4)

r = section_label(ws, r, "Truy vấn 1 — ‘bespoke ao dai saigon’", span=4)
rows1 = [
    ["🇻🇳 Việt Nam", "Local Pack: Minh Chau (5.0★/116), Kim Bespoke (4.9★/606), Byfas (4.1★/104). Organic: Maydo #2, Duan Tailor #6 (site riêng) + Tripadvisor/Kim Travel/TNK", "Có — nhiều", "Cạnh tranh thật. Thắng bằng Google Business Profile + review (vào Local Pack) và nội dung heritage sâu hơn product page"],
    ["🇺🇸 US", "Reddit #1, Kim Travel, LAHAVA, Facebook, Instagram, Tripadvisor; chỉ Mark&Vy (#7) là atelier", "Gần như không", "Góc “bespoke made in Saigon, thử trực tiếp” còn trống ở thị trường diaspora"],
]
r = table(ws, r, ["Địa điểm", "Ai đứng top trang 1", "Atelier Việt ở top?", "Ý nghĩa cho Maison"], rows1, widths=[14, 40, 18, 38], wrap_cols={2, 3, 4})
r = note_line(ws, r, "➡️ Truy vấn này khác hẳn theo địa điểm: ở VN atelier Việt đã rank (cạnh tranh + cần GBP); ở US gần như chỉ forum/seller (khoảng trống cho góc “made in Saigon”). PAA: “How much does an Ao Dai cost?” · “Can foreigners wear the Ao Dai?” · “Where to buy Ao Dai in HCMC?”", span=4, color="000000", italic=False, height=40)
r += 1

r = section_label(ws, r, "Truy vấn 2 — ‘wedding ao dai’", span=4)
rows2 = [
    ["🇻🇳 Việt Nam", "Nicole Bridal #1 (product) & #9 (blog) · LAHAVA #3 · AoDaiChiDau #7 · Pinterest/Etsy/The Knot", "Có", "Nhà cưới Việt rank được cả product lẫn blog → Maison làm tương tự được, cần nội dung tốt hơn"],
    ["🇺🇸 US", "Dream Dresses #1 · linh bridal #3 · East Meets Dress #5 · The Knot #6 · Amazon #7", "Không", "Seller diaspora + Amazon. Không có atelier Saigon → góc “made in Saigon” độc quyền"],
    ["🇦🇺 Úc", "Dream Dresses #1 · Le Vow Bridal #2 · Ao Dai Sydney #5 · Reddit · LAHAVA", "Không (có shop bản địa Úc)", "Cũng không có atelier Saigon, nhưng đã có shop Việt bản địa Úc → cạnh tranh diaspora cao hơn US"],
]
r = table(ws, r, ["Địa điểm", "Top organic trang 1", "Atelier/brand Việt ở top?", "Ý nghĩa cho Maison"], rows2, widths=[14, 40, 18, 38], wrap_cols={2, 3, 4})
r = note_line(ws, r, "➡️ Cụm “bespoke bridal áo dài made in Saigon” không bị atelier Saigon nào cạnh tranh ở thị trường diaspora US/AU (East Meets Dress/Dream Dresses bán theo số đo + ship, không có xưởng tại VN) → góc độc quyền cho Pillar B.", span=4, color="000000", italic=False, height=40)


# =====================================================================
# 4. SHORTLIST ƯU TIÊN
# =====================================================================
ws = new_sheet("4. Shortlist ưu tiên (Top 10)")
r = title_block(ws, "Shortlist Từ Khoá Ưu Tiên (đầu vào cho 10 blog brief)",
    "Xếp theo: khả năng lên top nhanh × giá trị conversion.", span=5)
sl = [
    [1, "wedding ao dai", "🏛 Pillar B", "Trends 50/50 (mạnh nhất); SERP US/AU = 0 atelier Saigon; VN có Nicole Bridal rank cả blog", "Từ mạnh nhất + góc ‘made in Saigon’ độc quyền ở thị trường diaspora"],
    [2, "traditional ao dai / ao dai meaning", "🏛/⚡", "Trends 57/50 (cao nhất nhóm info); related ‘what is ao dai’ rising +100%", "Phát hiện mới: nhóm info này còn mạnh hơn bridal → nam châm TOFU"],
    [3, "bespoke ao dai (saigon)", "🏛 Pillar A", "SW vol=0; SERP VN có atelier thật (Maydo, Duan) + Local Pack; SERP US gần như trống atelier", "Định vị lõi; ở VN cần GBP + nội dung sâu hơn product page"],
    [4, "how much does a custom ao dai cost", "⚡ Quick-win", "PAA #1 ở cả 2 truy vấn SERP đo được", "Câu hỏi top, intent mua, Maison trả lời = bắt khách"],
    [5, "how long to make an ao dai (pre-trip)", "⚡ Quick-win", "Khớp hành trình pre-trip (Proposal Nhóm 2)", "Phục vụ pre-trip funnel trực tiếp"],
    [6, "silk ao dai / chọn lụa", "⚡ Quick-win", "Trends 11/50, đỉnh tuần Dec'25=72", "Có cầu theo mùa + lợi thế phố lụa Lê Thánh Tôn"],
    [7, "custom ao dai vietnam (cho người nước ngoài)", "⚡ Quick-win", "PAA ‘Can foreigners wear the Ao Dai?’ xác minh", "Intent đặt may + expat/traveler"],
    [8, "evening wear saigon", "🏛 Pillar C", "Trends 1/50 nhưng đối thủ EN ≈ 0", "Occasion; trống hoàn toàn — rank dễ dù cầu nhỏ"],
    [9, "ao dai vs western wedding dress", "⚡ Quick-win", "PAA tea-ceremony đã xác minh (WebSearch)", "Kéo nhóm bridal về Pillar B"],
    [10, "modern ao dai", "⚡ Quick-win", "Related query ‘ao dai modern’=24", "Chủ đề có cầu, dễ rank"],
]
r = table(ws, r, ["Hạng", "Từ khoá chính", "Loại", "Số liệu hậu thuẫn", "Lý do ưu tiên"], sl, widths=[7, 38, 14, 50, 46], wrap_cols={2,4,5})
r = note_line(ws, r, "Lưu ý so với bản trước: ‘vietnamese couture’ đã rớt khỏi top ưu tiên (Trends thật 1/50, là blip chứ không phải xu hướng bền) — chỉ còn là góc định vị trong Pillar C. ‘traditional ao dai’ (57/50) được nâng lên hạng 2 sau khi đo số thật.", span=5, height=32)


# =====================================================================
# 5. TRENDS · GEO · RELATED
# =====================================================================
ws = new_sheet("5. Trends · Geo · Related")
r = title_block(ws, "Google Trends · Demand theo quốc gia · Từ khoá liên quan",
    "Reconfirm bằng Google Trends thật (SerpApi, 12 tháng, toàn cầu) — 11/6/2026. Trends là chỉ số tương đối 0–100, dùng để xếp hạng ưu tiên, không phải lượt tuyệt đối.", span=3)

r = section_label(ws, r, "5.1 Mức độ quan tâm thực (TIMESERIES, TB 12 tháng)", span=3)
t1 = [
    ["wedding ao dai", "50", "Vượt trội tuyệt đối — cầu ổn định quanh năm, đỉnh mạnh mùa Tết (T1–T2) & T4 (đạt 100 đầu T4/2026). Từ ‘thật’ duy nhất có volume đáng kể."],
    ["silk ao dai", "10", "Có thật nhưng ngắt quãng — nền bằng 0, thỉnh thoảng vọt (Dec'25=72, Feb'26=50). Theo mùa/dịp."],
    ["vietnamese couture", "1", "Gần như không có cầu ổn định. ‘↑33’ trong brief khách chỉ là blip ngắn hạn sau Couture Week, không bền."],
    ["bridal ao dai", "1", "Rất ít cầu dưới đúng cụm này — cầu thực nằm ở ‘wedding ao dai’."],
    ["bespoke ao dai", "0", "Không đo được cầu — xác nhận SimilarWeb. Là từ định vị thương hiệu, không phải từ tìm kiếm."],
]
r = table(ws, r, ["Từ khoá", "Điểm TB", "Diễn giải"], t1, widths=[22, 10, 80], wrap_cols={3})
r += 1

r = section_label(ws, r, "5.2 Demand đến từ đâu (GEO_MAP cho ‘wedding ao dai’)", span=3)
t2 = [
    ["🇦🇺 Australia", "100"],
    ["🇺🇸 United States", "100"],
    ["🇻🇳 Vietnam", "<1"],
    ["🇨🇦 Canada / 🇬🇧 UK / 🇩🇪 Germany", "<1"],
]
r = table(ws, r, ["Quốc gia", "Chỉ số"], t2, widths=[34, 12])
r = note_line(ws, r, "➡️ Xác nhận mạnh chiến lược tiếng Anh + đối tượng quốc tế. Người tìm ‘wedding ao dai’ là diaspora Việt ở Úc & Mỹ — gần như không ai ở VN gõ cụm tiếng Anh này. Nội dung bridal nên nói được với cô dâu gốc Việt ở Úc/Mỹ (đặt may từ xa rồi về VN/ship) — trùng khít pre-trip booking funnel.", span=3, color="000000", italic=False, height=34)
r += 1

r = section_label(ws, r, "5.3 Related queries thật cho head term ‘ao dai’ (đã lọc nhiễu)", span=3)
t3 = [
    ["Top (cầu ổn định)", "ao dai vietnam (98) · ao dai vietnamese (80) · ao dai dress (67) · traditional ao dai (34) · ao dai wedding (28) · ao dai modern (24) · vietnamese dress (24)"],
    ["Màu sắc (xác nhận bài #10)", "red ao dai (19) · ao dai trang / white (22) · black ao dai (rising +100%)"],
    ["Câu hỏi (xác nhận bài #10 là nam châm TOFU)", "what is ao dai (27, rising +100%)"],
    ["Dịp", "ao dai tet (18) · ao dai shop (17, intent mua)"],
    ["🔎 Phát hiện mới — ngách bị bỏ sót", "ao dai nam (43) & men ao dai (31) — áo dài nam/chú rể có cầu thật, khá cao. Maison có làm đồ nam/chú rể → bổ sung tín hiệu groom’s áo dài vào Pillar B (spoke mở rộng đợt sau)."],
]
r = table(ws, r, ["Nhóm", "Từ khoá (điểm)"], t3, widths=[40, 80], wrap_cols={2})


# =====================================================================
# 6. BẢN ĐỒ KHOẢNG TRỐNG
# =====================================================================
ws = new_sheet("6. Bản đồ khoảng trống")
r = title_block(ws, "Bản Đồ Khoảng Trống & Cơ Hội (Gap Map)",
    "Các khoảng trống thị trường đo được và cách Maison khai thác để lên top nhanh.", span=3)
gm = [
    ["Diaspora US/AU: “bespoke made in Saigon”", "SERP US/AU không có atelier Saigon (chỉ seller diaspora + Amazon); họ bán theo số đo + ship, không có xưởng VN", "Pillar B + bài bridal nhắm diaspora"],
    ["Thị trường VN: chiều sâu nội dung + GBP", "Atelier Việt đã rank (Maydo, Duan, Nicole Bridal) nhưng phần lớn là product page; Maison hơn ở craft + ảnh editorial + KOL. Cần Google Business Profile để vào Local Pack", "Pillar A + craftsmanship + GBP"],
    ["Pre-trip booking funnel", "Chưa ai tối ưu hành trình ‘đặt trước khi bay sang’ cho áo dài cao cấp", "CTA booking + bài ‘lên kế hoạch may áo dài khi đến Saigon’"],
    ["Occasion/evening wear Saigon (EN)", "Gần như trống hoàn toàn cho khách quốc tế", "Pillar C"],
    ["Occasion/couture (góc định vị)", "‘vietnamese couture’ Trends 12m chỉ 1/50 (blip, không bền) → không kỳ vọng traffic; dùng làm góc thương hiệu trong Pillar C", "Bài occasion gắn thương hiệu"],
    ["Lụa & thủ công (silk, embroidery)", "‘silk ao dai’ Trends 11/50 (đỉnh tuần Dec'25=72); Maison ở ngay phố lụa Lê Thánh Tôn (194)", "Bài ‘chọn lụa’ + craftsmanship"],
]
r = table(ws, r, ["Khoảng trống", "Vì sao Maison thắng nhanh", "Khai thác bằng"], gm, widths=[38, 52, 38], wrap_cols={1,2,3})


# =====================================================================
# 7. CÂU HỎI KHÁCH HỎI (PAA)
# =====================================================================
ws = new_sheet("7. Câu hỏi khách hỏi (PAA)")
r = title_block(ws, "Ngân Hàng Câu Hỏi Long-tail (People-Also-Ask) đã xác minh",
    "Lấy từ SERP thật + WebSearch — đây chính là nguyên liệu cho 7 bài quick-win. PAA = ‘People Also Ask’, hộp câu hỏi liên quan Google hiển thị.", span=2)
paa = [
    ["Chi phí", "“how much does a custom ao dai cost?” — đáp thực tế: từ ~300k VND (chợ) → $300–$1.000+ (designer)"],
    ["Thời gian", "“how long does it take to make an ao dai?” — 3–7 ngày; express 4–48h"],
    ["Lụa & vải", "Chợ vải Soai Kinh Lam / Tân Định; Nice Silk & Thu Silk trên Lê Thánh Tôn (cùng phố với Maison — 194 Lê Thánh Tôn)"],
    ["Ý nghĩa & màu sắc", "Đỏ = may mắn; dragon/phoenix = chú rể/cô dâu; khăn đóng"],
    ["Modern vs traditional (cách tân)", "Chủ đề đang hot"],
    ["Áo dài vs đầm cưới phương Tây", "Nhiều cô dâu mặc cả hai (tea ceremony + reception)"],
    ["Mặc cho dịp nào / cách mặc", "Tea ceremony, đám hỏi, lễ; fitting cho người nước ngoài"],
    ["Câu hỏi PAA bổ sung (từ truy vấn ‘bespoke ao dai saigon’)", "“How much does an Ao Dai cost in Vietnam?” · “Can foreigners wear the Ao Dai?” · “Where to buy Ao Dai in HCMC?” · “How much to get a custom dress in Vietnam?”"],
]
r = table(ws, r, ["Chủ đề câu hỏi", "Câu hỏi thật + đáp án xác minh"], paa, widths=[42, 86], wrap_cols={1,2})


# =====================================================================
# 8. 10 BLOG BRIEF
# =====================================================================
ws = new_sheet("8. 10 Blog Brief")
r = title_block(ws, "10 Blog Brief (mỗi bài 1 dòng)",
    "Strategic brief: định hướng từ khoá, intent, góc, internal link, meta. Nội dung chữ & kể chuyện do Maison cung cấp; Thiệu cấu trúc SEO & đăng. Ngôn ngữ bài: tiếng Anh. "
    "Vai trò: 🏛 Pillar (trụ) · ⚡ Quick-win. Phễu: TOFU (nhận biết) · MOFU (cân nhắc) · BOFU (sẵn sàng đặt).", span=13)
briefs = [
    [1, "🏛 Pillar A", "The Art of a Bespoke Áo Dài in Saigon",
     "bespoke ao dai saigon",
     "bespoke ao dai, tailored ao dai, custom ao dai vietnam, designer ao dai saigon, artisanal bespoke fashion, vietnamese dress tailor, made in saigon",
     "Trans + Info", "Nhóm 1 (expat) + Nhóm 2 (traveler)",
     "Định nghĩa ‘bespoke’ khác may sẵn/chợ; quy trình Maison (tư vấn→số đo→chọn lụa→thử→hoàn thiện); tay nghề thủ công; vì sao ‘made in Saigon’ có giá trị; ai nên đặt bespoke. Trang ‘tất cả đường dẫn về’ của cụm.",
     "→ #4, #6, #7, #9, /booking",
     "Bespoke Áo Dài in Saigon — Custom-Tailored by Maison Denude",
     "Discover the art of a bespoke áo dài, hand-tailored in Saigon. See the process, craftsmanship, and how to book your own custom fitting.",
     "1.500–2.200 từ", "MOFU→BOFU · ‘Book a bespoke consultation’"],
    [2, "🏛 Pillar B", "The Bridal Áo Dài: A Saigon Atelier's Guide to Your Wedding Áo Dài",
     "wedding ao dai",
     "bridal ao dai, ao dai wedding, ao dai bride, modern bridal ao dai, bespoke wedding saigon, silk ao dai bridal, wedding dress vietnam, embroidered ao dai, ao dai nam / men's ao dai (chú rể)",
     "Info → Trans", "Diaspora Việt ở 🇦🇺 Úc & 🇺🇸 Mỹ (AU=100, US=100, VN<1) + Nhóm 2",
     "Áo dài cưới là gì & ý nghĩa (đỏ/may mắn, dragon–phoenix, khăn đóng); truyền thống vs hiện đại; lụa & thêu tay; vì sao đặt may TẠI Saigon hơn ship từ nước ngoài (góc độc quyền vs East Meets Dress); gợi ý áo dài chú rể (ao dai nam=43); nói với cô dâu gốc Việt ở Úc/Mỹ; timeline đặt. Mùa vụ: đẩy trước cao điểm (T11–T1 Tết, T3 cưới xuân).",
     "→ #5, #9, #10, #6, /booking",
     "Bridal Áo Dài — Your Wedding Áo Dài, Made in Saigon",
     "A complete guide to the bridal áo dài: meaning, colors, fabrics, and how to have yours bespoke-tailored in Saigon for your wedding day.",
     "1.500–2.200 từ", "MOFU→BOFU · ‘Book a bridal áo dài fitting’"],
    [3, "🏛 Pillar C", "Occasion & Evening Wear in Saigon: Vietnamese Couture for Special Moments",
     "evening wear saigon / occasion wear vietnam",
     "evening dress ho chi minh, custom evening dress, evening gown vietnam, contemporary ao dai, vietnamese couture (brand-positioning, không kỳ vọng traffic), ao dai haute couture, vietnamese haute couture",
     "Trans + Brand", "Nhóm 1 (expat dự sự kiện) + Nhóm 2",
     "Đồ dự tiệc/sự kiện đặt may tại Saigon (trọng tâm intent); áo dài cách tân cho dịp trang trọng; dùng ‘vietnamese couture’ như góc định vị thương hiệu, KHÔNG coi là từ kéo traffic (Trends thật TB=1); editorial/KOL proof (Chi Pu, Văn Mai Hương); chọn đồ occasion thế nào.",
     "→ #8, #1, #6, /booking",
     "Evening & Occasion Wear in Saigon — Vietnamese Couture",
     "Bespoke evening and occasion wear in Saigon. Explore Vietnamese couture and custom-tailored looks for galas, events, and celebrations.",
     "1.300–1.800 từ", "MOFU→BOFU · ‘Book an occasion-wear consultation’"],
    [4, "⚡ Quick-win", "How Much Does a Bespoke Áo Dài Cost? (A Transparent Saigon Price Guide)",
     "how much does a custom ao dai cost",
     "ao dai price, custom ao dai cost saigon, bespoke ao dai price, ao dai tailoring cost",
     "Trans", "Nhóm 1 + 2",
     "Khung giá thật theo phân khúc (chợ → tầm trung → designer/bespoke); giá phụ thuộc gì (lụa, thêu tay, beading, độ phức tạp); vì sao bespoke cao cấp đáng giá; định vị minh bạch của Maison. Bắt người đang so giá rồi educate lên premium.",
     "→ #1, #6, /booking",
     "How Much Does a Bespoke Áo Dài Cost in Saigon?",
     "A transparent guide to bespoke áo dài pricing in Saigon — what affects the cost, and what to expect from a premium custom atelier.",
     "900–1.300 từ", "MOFU · ‘Get a personalised quote’"],
    [5, "⚡ Quick-win", "Getting an Áo Dài Made in Saigon: How Long It Takes & How to Plan Your Trip",
     "how long does it take to make an ao dai",
     "ao dai timeline, get ao dai made saigon, book ao dai before trip vietnam, ao dai fitting tourist",
     "Info → Trans", "Nhóm 2 (traveler — ‘đặt trước khi bay sang’)",
     "Timeline thực tế (số ngày, mấy lần thử); lời khuyên ‘đặt ngay đầu chuyến, đừng để sát ngày bay’; cách đặt trước online rồi đến thử (đúng funnel pre-trip); checklist chuẩn bị (số đo, ảnh tham khảo).",
     "→ #2, #1, #7, /booking",
     "How Long to Make an Áo Dài? Plan Your Saigon Visit",
     "How long does a bespoke áo dài take in Saigon? Timelines, fittings, and how to book before your trip so it's ready when you arrive.",
     "800–1.200 từ", "BOFU · ‘Reserve your fitting before you fly’"],
    [6, "⚡ Quick-win", "Choosing Silk for Your Áo Dài: A Guide from Saigon's Silk Street",
     "silk ao dai",
     "ao dai silk fabric, ao dai material, silk for ao dai, vietnamese silk, brocade ao dai",
     "Info → Trans", "Nhóm 1 + 2",
     "Các loại lụa/vải (lụa, brocade/gấm, chiffon…); lụa ảnh hưởng dáng & giá thế nào; lợi thế Maison ngay phố lụa Lê Thánh Tôn (194 — cùng phố Nice Silk/Thu Silk); chọn lụa theo dịp (cưới vs hằng ngày).",
     "→ #1, #4, #2, /booking",
     "Choosing Silk for Your Áo Dài — A Saigon Guide",
     "Silk, brocade, or chiffon? A guide to choosing the right fabric for your áo dài, from a bespoke atelier on Saigon's silk street.",
     "800–1.200 từ", "MOFU · ‘See our fabric selection in person’"],
    [7, "⚡ Quick-win", "A Foreigner's Guide to Having a Custom Áo Dài Made in Vietnam",
     "custom ao dai vietnam",
     "ao dai for foreigners, custom ao dai for tourists, where to get ao dai made district 1, tailored ao dai saigon",
     "Trans", "Nhóm 1 + 2 (người nước ngoài lần đầu)",
     "Quy trình từ góc người nước ngoài: đặt lịch, giao tiếp số đo, chọn dáng hợp người phương Tây, etiquette mặc áo dài, thử & chỉnh; gỡ rào cản ‘tôi không phải người Việt mặc có hợp không’; vì sao chọn atelier cao cấp thay vì quầy chợ.",
     "→ #1, #5, #10, /booking",
     "Custom Áo Dài in Vietnam — A Foreigner's Guide",
     "New to the áo dài? A foreigner's guide to having one custom-made in Saigon: the process, fit, etiquette, and how to book.",
     "900–1.300 từ", "MOFU→BOFU · ‘Book your first fitting’"],
    [8, "⚡ Quick-win", "Modern vs Traditional Áo Dài: Which Style Is Right for You?",
     "modern ao dai / contemporary ao dai",
     "cach tan ao dai, modern vs traditional ao dai, contemporary ao dai, ao dai styles",
     "Info", "Nhóm 1 + 2",
     "So sánh truyền thống (cổ cao, dáng cổ điển, lụa/gấm) vs cách tân (dáng ngắn, cổ/tay biến tấu, vải mới); chọn style theo dịp & cá tính; cách Maison cân bằng heritage × sultry/feminine power. Cầu nối sang occasion (Pillar C).",
     "→ #3, #1, #2, /booking",
     "Modern vs Traditional Áo Dài — Which Style Suits You?",
     "Modern or traditional áo dài? Compare the styles, collars, fabrics, and occasions to find the silhouette that's right for you.",
     "800–1.200 từ", "TOFU→MOFU · ‘Design your own áo dài with us’"],
    [9, "⚡ Quick-win", "Áo Dài vs Western Wedding Dress: Why Brides Choose Both",
     "ao dai vs western wedding dress",
     "vietnamese wedding dress, tea ceremony dress, ao dai tea ceremony, bridesmaids ao dai, fusion wedding ao dai",
     "Info", "Nhóm 2 + diaspora",
     "Vì sao nhiều cô dâu mặc cả hai (áo dài cho tea ceremony/đám hỏi + đầm trắng reception); etiquette màu (đỏ/trắng cho cô dâu, khách tránh đỏ/trắng/đen); fusion design; gợi ý đặt may áo dài cưới + áo dài cho dàn bưng quả.",
     "→ #2, #10, #6, /booking",
     "Áo Dài vs Western Wedding Dress — Why Brides Wear Both",
     "Áo dài or white gown? Why many brides wear both, what to wear to the tea ceremony, and how to plan your bridal áo dài.",
     "800–1.200 từ", "TOFU→MOFU · ‘Plan your bridal áo dài’"],
    [10, "⚡ Quick-win", "The Meaning of the Áo Dài: Colors, Symbols & How to Wear It",
     "ao dai meaning (+ how to wear ao dai, ao dai colors)",
     "what is ao dai, ao dai symbolism, ao dai colors meaning, heritage ao dai, traditional ao dai, khan dong",
     "Info (TOFU rộng nhất)", "Tất cả (cửa ngõ nhận biết)",
     "Áo dài là gì, lịch sử ngắn; ý nghĩa màu (đỏ/may mắn, trắng, gold, pastel); biểu tượng dragon–phoenix; khăn đóng; cách mặc đúng dịp. Bài ‘nam châm’ thu hút top-of-funnel rồi điều hướng xuống bespoke/bridal.",
     "→ #1, #2, #8, /booking",
     "The Meaning of the Áo Dài — Colors, Symbols & How to Wear",
     "What does the áo dài mean? A guide to its colors, symbols, history, and how to wear Vietnam's national dress for every occasion.",
     "900–1.300 từ", "TOFU · ‘Ready for your own? Book a consultation’"],
]
hdrb = ["#", "Vai trò", "Tên bài", "Từ khoá chính", "Từ khoá phụ / semantic", "Intent", "Đối tượng",
        "Góc tiếp cận / ý chính", "Internal links", "Meta title (≤60)", "Meta description (≤155)", "Độ dài", "Phễu / CTA"]
r = table(ws, r, hdrb, briefs,
          widths=[4, 11, 30, 26, 34, 13, 24, 56, 20, 34, 40, 13, 26],
          wrap_cols={3,4,5,7,8,9,10,11,13})


# =====================================================================
# 9. KIẾN TRÚC CỤM CHỦ ĐỀ
# =====================================================================
ws = new_sheet("9. Kiến trúc cụm chủ đề")
r = title_block(ws, "Kiến Trúc Cụm Chủ Đề (Topic Cluster) — vì sao thiết kế thế này",
    "KPI: Top-10 cho ≥3 từ khoá trong 3 tháng. Site mới chưa có uy tín miền (domain authority) → không thể ‘đấm’ 1 từ khoá lớn. "
    "Dùng mô hình hub-and-spoke (trục–nan hoa): trụ rộng gom uy tín, nhánh long-tail lên top nhanh và link ngược về trụ.", span=2)

r = section_label(ws, r, "Logic mô hình", span=2)
for t in [
    "• 3 bài trụ (Pillar) = hub rộng, gom authority, là trang Maison muốn rank cho từ khoá định vị.",
    "• 7 bài hỗ trợ (Quick-win) = spoke long-tail, lên top nhanh từng bài (cạnh tranh thấp), mỗi bài link ngược về pillar → dồn tín hiệu liên quan cho hub.",
    "• Kết quả: Google đọc được Maison là chuyên gia về ‘bespoke áo dài Saigon’ — semantic network giúp pillar rank bền, quick-win rank nhanh.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, t); c.font = F(10); c.alignment = WRAP; ws.row_dimensions[r].height = 30
    r += 1
r += 1

r = section_label(ws, r, "Sơ đồ cụm (mọi bài đều CTA về /booking)", span=2)
diagram = (
"                         ┌─────────────────────────────┐\n"
"                         │   /booking  (đích chuyển đổi) │\n"
"                         └──────────────▲──────────────┘\n"
"        mọi bài đều CTA về booking      │\n"
"   ┌────────────────┐   ┌───────────────┴────────────┐   ┌────────────────┐\n"
"   │ PILLAR A        │   │ PILLAR B                    │   │ PILLAR C        │\n"
"   │ Bespoke Áo Dài  │   │ Wedding / Bridal Áo Dài     │   │ Occasion &      │\n"
"   │ (Saigon)        │   │ (Made in Saigon)            │   │ Couture Saigon  │\n"
"   └───▲───▲───▲─────┘   └────▲────────▲────────▲──────┘   └─────▲────────▲──┘\n"
"       │   │   │              │        │        │                │        │\n"
"      #4  #6  #7             #9       #5       #10              #8    (#10 dùng chung)\n"
"   (chi phí)(lụa)(custom)  (vs western)(pre-trip)(meaning)    (modern)\n"
)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(r, 1, diagram); c.font = Font(name="Consolas", size=9); c.alignment = WRAP
ws.row_dimensions[r].height = 190
r += 2

r = section_label(ws, r, "Quy tắc liên kết nội bộ (bắt buộc, để không có bài mồ côi)", span=2)
rules = [
    ["Mỗi Quick-win link tới", "(a) ít nhất 1 Pillar mẹ, (b) /booking, (c) ≥1 quick-win anh em liên quan."],
    ["Mỗi Pillar link tới", "(a) ≥2 quick-win con, (b) /booking, (c) ≥1 pillar khác."],
    ["Anchor text", "Dùng từ khoá tự nhiên (vd: ‘see how a bespoke áo dài is made’), không nhồi."],
]
r = table(ws, r, ["Đối tượng", "Quy tắc"], rules, widths=[26, 86], wrap_cols={2})
ws.column_dimensions["A"].width = 26


# =====================================================================
# 10. LỊCH XUẤT BẢN
# =====================================================================
ws = new_sheet("10. Lịch xuất bản")
r = title_block(ws, "Lịch Xuất Bản Đề Xuất (để rank nhanh & bồi cụm hợp lý)",
    "Đăng quick-win trước (rank nhanh, có data sớm), pillar xen kẽ để có đích internal link. Khớp với plan: blog content nhận từ Maison ở T4, đăng & tối ưu T5–T6.", span=3)
sched = [
    ["1", "#10 (meaning) + #1 (Pillar A)", "TOFU nam châm + trụ lõi làm đích link"],
    ["2", "#4 (cost) + #7 (foreigner guide) + #2 (Pillar B)", "Bắt intent mua sớm + trụ bridal"],
    ["3", "#5 (timeline) + #6 (silk) + #9 (vs western)", "Funnel pre-trip + bridal spokes"],
    ["4", "#8 (modern) + #3 (Pillar C)", "Đóng cụm occasion/couture"],
]
r = table(ws, r, ["Đợt", "Bài", "Lý do"], sched, widths=[8, 50, 50], wrap_cols={2,3})


# =====================================================================
# 11. CHECKLIST SEO KỸ THUẬT
# =====================================================================
ws = new_sheet("11. Checklist SEO kỹ thuật")
r = title_block(ws, "Yêu Cầu SEO Kỹ Thuật Cho Mọi Bài (checklist khi đăng)",
    "Dành cho đội kỹ thuật/đăng tải. Tích vào ô khi hoàn thành.", span=2)
checks = [
    "Từ khoá chính xuất hiện ở title, H1, đoạn mở đầu, ≥1 H2, URL slug (vd /blog/bespoke-ao-dai-saigon).",
    "1 H1 duy nhất; H2/H3 dùng từ khoá phụ & câu hỏi PAA.",
    "Meta title ≤60 ký tự, meta description ≤155 ký tự (đã soạn sẵn ở mỗi brief).",
    "Internal link đúng quy tắc hub-and-spoke (sheet 9) — không bài mồ côi.",
    "Schema BlogPosting + Open Graph + Twitter Card (theo plan.md Phase 6).",
    "Ảnh WebP, nén, alt text mô tả có từ khoá (Maison có sẵn ảnh KOL/editorial).",
    "CTA /booking rõ ràng trong bài (phục vụ conversion + pre-trip funnel).",
    "Dùng ‘ao dai’ (không dấu) trong slug/từ khoá kỹ thuật; ‘áo dài’ có dấu trong nội dung biên tập.",
    "Nhắc địa lý ‘Saigon / Ho Chi Minh City / District 1 / 194 Lê Thánh Tôn’ để bồi local SEO + LocalBusiness entity.",
]
crows = [["☐", c] for c in checks]
r = table(ws, r, ["✓", "Mục kiểm tra"], crows, widths=[6, 110], wrap_cols={2})
ws.column_dimensions["B"].width = 110


# =====================================================================
# 12. PHƯƠNG PHÁP & ĐỘ TIN CẬY
# =====================================================================
ws = new_sheet("12. Phương pháp & Độ tin cậy")
r = title_block(ws, "Lưu Ý Phương Pháp & Độ Tin Cậy Dữ Liệu",
    "Toàn bộ số liệu thô lưu tại data_logs/ để audit & không phải fetch lại.", span=2)
method = [
    ["SimilarWeb volume mỏng = bình thường & là tín hiệu tốt", "Nhiều từ trả ‘≈0 đo được’ — bình thường với ngách siêu chuyên + tiếng Anh + địa phương VN, và là tín hiệu tốt cho mục tiêu lên top nhanh. Volume head term ao dai (~62–72K, 100% info) là mốc neo đáng tin."],
    ["Cạnh tranh đánh giá bằng SERP đo thật (đa địa điểm)", "SimilarWeb trả difficulty/competition = null → đánh giá cạnh tranh dựa trên vị trí SERP đo thật theo ma trận VN·US·AU (SerpApi, 11/6/2026). SERP khác nhau giữa thị trường nội địa và diaspora nên cần đo nhiều nơi."],
    ["Chuẩn hoá Trends xuyên batch", "Google Trends chỉ chuẩn hoá nội bộ mỗi nhóm ≤5 từ. Để so cả 34 từ trên một thang chung, mỗi batch đều chứa ‘wedding ao dai’ làm neo, rồi nhân lại theo công thức: điểm = avg_batch × 50/anchor_batch. 16/34 từ chấm trực tiếp; còn lại suy theo pattern (‘~0 ước’)."],
    ["Bộ Trends nội bộ của khách KHÁC thang", "Bộ khách (wedding ao dai=80, silk=47, couture=33, bridal=27) là điểm so sánh nội bộ một batch, không cùng thang với điểm chuẩn hoá ở đây. Thứ tự đúng một phần (wedding ao dai dẫn đầu) nhưng couture bị thổi phồng (thật=1, không phải 33)."],
    ["Cách đọc Google Trends", "Trends là chỉ số tương đối 0–100, không phải volume tuyệt đối — dùng để xếp hạng ưu tiên, không quy ra số lượt. wedding ao dai mang tính mùa vụ rất rõ (đỉnh Tết & mùa cưới) → đẩy nội dung trước cao điểm (T11–T1 Tết, T3 cưới xuân)."],
    ["Cập nhật sau go-live", "Khi GSC (Search Console) của site mới chạy, thay các ước lượng này bằng impression/click thực tế và cập nhật lại shortlist (mốc Tuần 7: ‘Bản đồ từ khoá + xếp hạng cơ sở’)."],
]
r = table(ws, r, ["Chủ đề", "Lưu ý"], method, widths=[40, 86], wrap_cols={1,2})


# =====================================================================
# 13. BÀN GIAO & GHI CHÚ
# =====================================================================
ws = new_sheet("13. Bàn giao & Ghi chú")
r = title_block(ws, "Bàn Giao & Liên Kết Công Việc",
    "Liên kết với công việc của Maddy, thay thế bảng cũ, và lộ trình cập nhật.", span=2)

r = section_label(ws, r, "Bàn giao & liên kết với Maddy", span=2)
hand = [
    "10 blog brief (sheet 8) là đầu vào cho mốc T3 của Maddy (‘Gửi cuốn chiếu blog theo brief Thiệu cấp ở T1’).",
    "Góc ‘3 thị trường trọng điểm’ của Maddy có thể mở rộng cụm nội dung sau (vd thêm bài Dubai/Singapore/Hàn) — nhưng KHÔNG thuộc 10 brief nền tảng này; 10 brief tập trung Nhóm 1+2 để đạt KPI nhanh trước.",
    "Reconfirm Google Trends: demand ‘wedding ao dai’ đến gần như 100% từ Úc + Mỹ (diaspora), Việt Nam <1. → Maddy nên cân nhắc Úc là một thị trường diaspora tiềm năng bên cạnh Mỹ.",
]
for t in hand:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, "• " + t); c.font = F(10); c.alignment = WRAP; ws.row_dimensions[r].height = 30
    r += 1
r += 1

r = section_label(ws, r, "Thay thế bảng cũ ‘SEO Blog Strategy’ (5 chủ đề trong management/plan.md)", span=2)
absorb = [
    ["Story behind bespoke áo dài", "#1 (Pillar A)"],
    ["Maison at major event", "#3 (Pillar C) — editorial/KOL angle"],
    ["Choose occasion wear", "#3 + #8"],
    ["Vietnamese fashion on world map", "#3 (couture angle)"],
    ["Bespoke vs ready-to-wear", "#4 + #7"],
]
r = table(ws, r, ["Chủ đề cũ (5 topic)", "Hấp thụ vào brief mới"], absorb, widths=[44, 44], wrap_cols={1,2})
r += 1
r = note_line(ws, r, "Cập nhật sau go-live: khi Search Console có data thật (Tuần 7+), đối chiếu impression/click thực tế → ưu tiên viết tiếp/đào sâu các bài đang lên top, bồi thêm spoke cho pillar nào rank tốt nhất.", span=2, color="000000", italic=False, height=30)


# =====================================================================
# 14. THUẬT NGỮ (GLOSSARY)
# =====================================================================
ws = new_sheet("14. Thuật ngữ (Glossary)")
r = title_block(ws, "Thuật Ngữ SEO — Giải Nghĩa Dễ Hiểu",
    "Bảng tra cứu nhanh các thuật ngữ dùng trong tài liệu này.", span=2)
gloss = [
    ["SEO", "Tối ưu hoá để website xuất hiện cao trên kết quả tìm kiếm Google (miễn phí, không phải quảng cáo)."],
    ["Head term (từ khoá lớn)", "Từ khoá ngắn, rất nhiều người tìm, cạnh tranh cao (vd ‘ao dai’). Khó lên top với site mới."],
    ["Long-tail (từ khoá dài)", "Cụm từ dài, cụ thể, ít người tìm hơn nhưng ý định rõ và cạnh tranh thấp (vd ‘bespoke ao dai saigon’) → dễ lên top nhanh."],
    ["Search intent (ý định tìm kiếm)", "Lý do người ta gõ từ khoá. Info = tìm hiểu · Trans (transactional) = muốn mua/đặt · Local = tìm chỗ gần · Brand = tìm thương hiệu cụ thể."],
    ["Volume (lượng tìm kiếm)", "Số lượt tìm một từ khoá mỗi tháng. Ở đây lấy từ SimilarWeb."],
    ["Google Trends (0–100)", "Chỉ số tương đối mức độ quan tâm theo thời gian — 100 = đỉnh, 0 = rất nhỏ so với đỉnh. Dùng để xếp hạng ưu tiên, KHÔNG phải số lượt tuyệt đối."],
    ["SERP", "Search Engine Results Page — trang kết quả Google. ‘Đo SERP’ = xem thực tế ai đang đứng trang 1."],
    ["AI Overview", "Hộp tóm tắt do AI của Google tạo, hiện trên đầu một số kết quả tìm kiếm."],
    ["PAA (People Also Ask)", "Hộp ‘Mọi người cũng hỏi’ trên Google — nguồn câu hỏi thật để viết bài."],
    ["Pillar (bài trụ)", "Bài dài, bao quát một chủ đề lớn, là ‘hub’ trung tâm gom uy tín cho cả cụm."],
    ["Quick-win / Spoke (bài nhánh)", "Bài ngắn nhắm từ khoá dài cụ thể, lên top nhanh, link ngược về bài trụ."],
    ["Hub-and-spoke (trục–nan hoa)", "Cách tổ chức nội dung: 1 trụ trung tâm + nhiều nhánh xoay quanh, liên kết với nhau."],
    ["Semantic network (mạng ngữ nghĩa)", "Tập hợp bài viết liên quan, liên kết nội bộ, giúp Google hiểu site là chuyên gia về chủ đề đó."],
    ["TOFU / MOFU / BOFU", "Các tầng phễu mua hàng: TOFU = nhận biết (mới biết đến) · MOFU = cân nhắc · BOFU = sẵn sàng đặt."],
    ["Funnel (phễu)", "Hành trình từ lúc khách biết đến thương hiệu → cân nhắc → quyết định mua."],
    ["Pre-trip booking funnel", "Hành trình khách quốc tế đặt may áo dài TRƯỚC khi bay sang VN, đến nơi chỉ việc thử & lấy."],
    ["CTA (Call To Action)", "Lời kêu gọi hành động trong bài, vd nút ‘Đặt lịch tư vấn’."],
    ["Meta title / Meta description", "Tiêu đề & đoạn mô tả ngắn hiển thị trên kết quả Google (title ≤60 ký tự, description ≤155)."],
    ["Internal link (liên kết nội bộ)", "Link giữa các bài trong cùng website — giúp giữ khách & dồn tín hiệu SEO."],
    ["Anchor text", "Đoạn chữ được gắn link (vd ‘bespoke áo dài’) — nên dùng từ khoá tự nhiên."],
    ["Slug / URL", "Phần đuôi địa chỉ bài viết, vd /blog/bespoke-ao-dai-saigon."],
    ["Schema BlogPosting", "Đoạn mã đánh dấu giúp Google hiểu trang là một bài blog (rich result)."],
    ["E-E-A-T", "Tiêu chí chất lượng Google: Experience–Expertise–Authoritativeness–Trust (kinh nghiệm–chuyên môn–thẩm quyền–độ tin cậy)."],
    ["Diaspora", "Cộng đồng người Việt sống ở nước ngoài (ở đây: Úc & Mỹ) — nhóm tìm ‘wedding ao dai’ nhiều nhất."],
    ["Bespoke", "May đo riêng hoàn toàn theo khách (khác ‘ready-to-wear’ may sẵn). Phân khúc cao cấp."],
    ["Atelier", "Nhà may/xưởng thời trang cao cấp."],
    ["GSC (Google Search Console)", "Công cụ miễn phí của Google cho biết site thật sự hiển thị & được click với từ khoá nào — dùng để cập nhật chiến lược sau go-live."],
    ["UGC", "User-Generated Content — nội dung do người dùng tạo (forum, review, mạng xã hội)."],
    ["RTW (ready-to-wear)", "Đồ may sẵn theo size, không may đo riêng."],
]
r = table(ws, r, ["Thuật ngữ", "Giải nghĩa (tiếng Việt)"], gloss, widths=[34, 92], wrap_cols={1,2})


wb.save(OUT)
print("SAVED:", OUT)
print("SHEETS:", wb.sheetnames)
